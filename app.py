from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse
import json, sqlite3, threading, webbrowser, math, os, subprocess
from datetime import datetime, timedelta
from sentos_client import SentosClient, SentosError, product_list
from secure_store import save_credentials, load_credentials
from urllib.request import urlopen

ROOT = Path(__file__).parent.resolve()
DB = ROOT / "fiyatlar.db"
VERSION = (ROOT / 'VERSION').read_text().strip() if (ROOT / 'VERSION').exists() else '0.0.0'
UPDATE_VERSION_URL = 'https://raw.githubusercontent.com/Kochleroguz/kochler-fiyat-otomasyonu/main/VERSION'

CHANNELS = {
    "trendyol": ("Trendyol", 5.99, 0.0),
    "hepsiburada": ("Hepsiburada", 12.60, 0.008),
    "n11": ("N11", 0.0, 0.0205),
    "pazarama": ("Pazarama", 13.08, 0.0),
    "amazon": ("Amazon", 0.0, 0.0),
    "idefix": ("İdefix", 12.0, 0.0),
    "boyner": ("Boyner", 12.0, 0.0),
    "temu": ("Temu", 0.0, 0.0),
    "web": ("Web Sitesi", 0.0, 0.0),
    "perakende": ("Perakende", 0.0, 0.0),
}

def num(v):
    if v is None or v == "": return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("₺", "").replace("TL", "").replace(" ", "")
    if "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
    elif "," in s: s = s.replace(",", ".")
    try: return float(s)
    except: return 0.0

def db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    with db() as c:
        c.executescript('''
        CREATE TABLE IF NOT EXISTS products(id INTEGER PRIMARY KEY, sku TEXT UNIQUE, name TEXT, barcode TEXT, category TEXT, cost REAL, web_price REAL, shipping REAL, min_margin REAL DEFAULT .02);
        CREATE TABLE IF NOT EXISTS channels(code TEXT PRIMARY KEY, name TEXT, fixed_fee REAL, extra_rate REAL);
        CREATE TABLE IF NOT EXISTS commissions(product_id INTEGER, channel TEXT, rate REAL, PRIMARY KEY(product_id,channel));
        CREATE TABLE IF NOT EXISTS category_commissions(category TEXT, channel TEXT, rate REAL, updated_at TEXT DEFAULT CURRENT_TIMESTAMP, PRIMARY KEY(category,channel));
        CREATE TABLE IF NOT EXISTS prices(product_id INTEGER, channel TEXT, price REAL, source TEXT DEFAULT 'excel', updated_at TEXT DEFAULT CURRENT_TIMESTAMP, PRIMARY KEY(product_id,channel));
        CREATE TABLE IF NOT EXISTS price_history(id INTEGER PRIMARY KEY, product_id INTEGER, channel TEXT, old_price REAL, new_price REAL, created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS offers(id INTEGER PRIMARY KEY, source TEXT, product_id INTEGER, source_sku TEXT, barcode TEXT, band INTEGER, low REAL, high REAL, rate REAL, starts TEXT, ends TEXT, status TEXT DEFAULT 'taslak');
        CREATE TABLE IF NOT EXISTS campaigns(id INTEGER PRIMARY KEY, product_id INTEGER, channel TEXT, normal_price REAL, campaign_price REAL, normal_rate REAL, campaign_rate REAL, starts TEXT, ends TEXT, seller_discount REAL DEFAULT 0, marketplace_share REAL DEFAULT 0, status TEXT DEFAULT 'taslak', created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS aliases(channel TEXT, external_sku TEXT, product_id INTEGER, PRIMARY KEY(channel,external_sku));
        CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT);
        CREATE TABLE IF NOT EXISTS shipping_rates(desi INTEGER PRIMARY KEY, price REAL, updated_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS sentos_products(sentos_id INTEGER PRIMARY KEY, sku TEXT, name TEXT, barcode TEXT, purchase_price REAL, vat_rate REAL, desi REAL, prices_json TEXT, synced_at TEXT);
        CREATE TABLE IF NOT EXISTS sync_log(id INTEGER PRIMARY KEY, kind TEXT, status TEXT, detail TEXT, item_count INTEGER DEFAULT 0, created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS product_events(id INTEGER PRIMARY KEY, product_id INTEGER, event_type TEXT NOT NULL, detail TEXT, created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS price_update_runs(id INTEGER PRIMARY KEY, status TEXT NOT NULL DEFAULT 'taslak', requested_count INTEGER DEFAULT 0, sent_count INTEGER DEFAULT 0, failed_count INTEGER DEFAULT 0, detail TEXT, created_at TEXT DEFAULT CURRENT_TIMESTAMP, completed_at TEXT);
        CREATE TABLE IF NOT EXISTS price_update_items(id INTEGER PRIMARY KEY, run_id INTEGER, product_id INTEGER, channel TEXT, old_price REAL, new_price REAL, status TEXT DEFAULT 'taslak', sentos_message TEXT, verified_price REAL, created_at TEXT DEFAULT CURRENT_TIMESTAMP, updated_at TEXT DEFAULT CURRENT_TIMESTAMP);
        ''')
        cols={x[1] for x in c.execute("PRAGMA table_info(products)")}
        additions={
            'desi':'INTEGER DEFAULT 0',
            'sentos_id':'INTEGER',
            'sentos_sku':'TEXT',
            'vat_rate':'REAL DEFAULT 20',
            'match_status':"TEXT DEFAULT 'bekliyor'",
            'match_note':'TEXT',
            'last_synced_at':'TEXT',
            'web_price_updated_at':'TEXT',
            'last_price_send_status':"TEXT DEFAULT 'gonderilmedi'",
            'last_price_sent_at':'TEXT'
        }
        for name, definition in additions.items():
            if name not in cols: c.execute(f"ALTER TABLE products ADD COLUMN {name} {definition}")
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS products_sentos_id_unique ON products(sentos_id) WHERE sentos_id IS NOT NULL")
        for code,(name,fixed,extra) in CHANNELS.items(): c.execute("INSERT OR IGNORE INTO channels VALUES(?,?,?,?)",(code,name,fixed,extra))
        c.execute("INSERT OR IGNORE INTO settings VALUES('sentos_mode','demo')")

def normalize_category_commissions():
    with db() as c:
        if c.execute("SELECT 1 FROM category_commissions LIMIT 1").fetchone(): return
        rows=c.execute('''SELECT p.category,cm.channel,cm.rate,COUNT(*) n FROM commissions cm JOIN products p ON p.id=cm.product_id WHERE p.category<>'' AND cm.channel NOT IN ('web','perakende') GROUP BY p.category,cm.channel,cm.rate ORDER BY n DESC''').fetchall()
        chosen={}
        for r in rows: chosen.setdefault((r['category'],r['channel']),r['rate'])
        for (category,channel),rate in chosen.items():
            c.execute("INSERT INTO category_commissions(category,channel,rate) VALUES(?,?,?)",(category,channel,rate))
            c.execute("DELETE FROM commissions WHERE channel=? AND rate=? AND product_id IN (SELECT id FROM products WHERE category=?)",(channel,rate,category))

def import_master(path):
    from openpyxl import load_workbook
    wf = load_workbook(path, data_only=True, read_only=True)
    s, cs, ks = wf['Fiyat Listesi'], wf['Ürün Komisyon Listesi'], wf['Kargo Fiyatları']
    shipping = {int(num(r[0])):num(r[1]) for r in ks.iter_rows(min_row=2, values_only=True) if r[0] is not None}
    comm = {}
    for r in cs.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        comm[str(r[0]).strip()] = dict(zip(['trendyol','hepsiburada','n11','pazarama','amazon','idefix','boyner','temu'], [num(r[i]) for i in [3,5,7,9,11,13,15,17]]))
    count=0
    with db() as c:
        for desi,price in shipping.items(): c.execute("INSERT OR REPLACE INTO shipping_rates(desi,price) VALUES(?,?)",(desi,price))
        for r in s.iter_rows(min_row=2, values_only=True):
            if not r[0]: continue
            sku=str(r[0]).strip(); desi=int(num(r[5])); ship=shipping.get(desi,0)
            c.execute("INSERT INTO products(sku,name,barcode,category,cost,web_price,shipping,min_margin,desi) VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(sku) DO UPDATE SET name=excluded.name,barcode=excluded.barcode,category=excluded.category,cost=excluded.cost,web_price=excluded.web_price,shipping=excluded.shipping,desi=excluded.desi",(sku,str(r[1] or ''),str(r[2] or ''),str(r[3] or ''),num(r[7]),num(r[8]),ship,num(r[9]) or .02,desi))
            pid=c.execute("SELECT id FROM products WHERE sku=?",(sku,)).fetchone()[0]
            for ch,rate in comm.get(sku,{}).items(): c.execute("INSERT OR REPLACE INTO commissions VALUES(?,?,?)",(pid,ch,rate))
            c.execute("INSERT OR REPLACE INTO commissions VALUES(?,?,?)",(pid,'web',num(r[9])))
            c.execute("INSERT OR IGNORE INTO commissions VALUES(?,?,?)",(pid,'perakende',0))
            count+=1
        price_columns = {'web':8,'perakende':8,'trendyol':13,'hepsiburada':16,'n11':18,'pazarama':21,'amazon':24,'idefix':26,'boyner':28,'temu':30}
        for r in s.iter_rows(min_row=2, values_only=True):
            if not r[0]: continue
            x=c.execute("SELECT id FROM products WHERE sku=?",(str(r[0]).strip(),)).fetchone()
            if not x: continue
            for ch,col in price_columns.items():
                price=num(r[col])
                if price>0: c.execute("INSERT OR REPLACE INTO prices(product_id,channel,price) VALUES(?,?,?)",(x[0],ch,price))
    return count

def match_product(c, sku, barcode, channel):
    if barcode:
        x=c.execute("SELECT id FROM products WHERE barcode=?",(str(barcode).strip(),)).fetchone()
        if x:return x[0]
    if sku:
        x=c.execute("SELECT id FROM products WHERE upper(sku)=upper(?)",(str(sku).strip(),)).fetchone()
        if x:return x[0]
        x=c.execute("SELECT product_id FROM aliases WHERE channel=? AND upper(external_sku)=upper(?)",(channel,str(sku).strip())).fetchone()
        if x:return x[0]
    return None

def import_offers(path):
    source='trendyol' if str(path).lower().endswith('xlsx') else 'n11'
    rows=[]
    if source=='trendyol':
        from openpyxl import load_workbook
        s=load_workbook(path,data_only=True,read_only=True).active
        for r in s.iter_rows(min_row=2,values_only=True):
            if not r[0]: continue
            for window,date_col in [(0,14),(1,19)]:
                dates=str(r[date_col] or '')
                for band in range(4):
                    low=num(r[8+band*2] if band==0 else r[10+(band-1)*2]); high=num(r[9+band*2]) if band<3 else 10**12
                    rate=num(r[15+window*5+band]);
                    if rate: rows.append((source,r[2],r[1],band+1,low,high,rate,dates,dates))
    else:
        import xlrd
        s=xlrd.open_workbook(path).sheet_by_index(0)
        for i in range(11,s.nrows):
            r=s.row_values(i)
            if not r[0]: continue
            start = xlrd.xldate_as_datetime(r[14],0).isoformat() if isinstance(r[14],(int,float)) and r[14] else ''
            end = xlrd.xldate_as_datetime(r[15],0).isoformat() if isinstance(r[15],(int,float)) and r[15] else ''
            for band in range(3):
                j=5+band*3; rows.append((source,r[1],r[4],band+1,num(r[j+1]),num(r[j]),num(r[j+2]),start,end))
    matched=0
    with db() as c:
        c.execute("DELETE FROM offers WHERE source=?",(source,))
        for src,sku,barcode,band,low,high,rate,starts,ends in rows:
            pid=match_product(c,sku,barcode,src); matched += bool(pid)
            c.execute("INSERT INTO offers(source,product_id,source_sku,barcode,band,low,high,rate,starts,ends) VALUES(?,?,?,?,?,?,?,?,?,?)",(src,pid,str(sku or ''),str(barcode or ''),band,low,high,rate/100 if rate>1 else rate,starts,ends))
    return {'source':source,'rows':len(rows),'matched':matched,'unmatched':len(rows)-matched}

def payload():
    with db() as c:
        products=[dict(x) for x in c.execute('''SELECT p.*, COUNT(DISTINCT o.source||o.source_sku) offer_count FROM products p LEFT JOIN offers o ON o.product_id=p.id GROUP BY p.id ORDER BY p.sku''')]
        channels=[dict(x) for x in c.execute("SELECT * FROM channels ORDER BY name")]
        for ch in channels:
            ch['vat_exclusive'] = ch['code']=='temu'
            ch['vat_rate'] = .20 if ch['code']=='temu' else 0
        offers=[dict(x) for x in c.execute('''SELECT o.*,p.sku,p.name FROM offers o LEFT JOIN products p ON p.id=o.product_id ORDER BY o.source,o.source_sku,o.band''')]
        campaigns=[dict(x) for x in c.execute('''SELECT c.*,p.sku,p.name,ch.name channel_name FROM campaigns c JOIN products p ON p.id=c.product_id JOIN channels ch ON ch.code=c.channel ORDER BY c.starts''')]
        category_commissions={}
        for x in c.execute('SELECT category,channel,rate FROM category_commissions'): category_commissions.setdefault(x[0],{})[x[1]]=x[2]
        commission_overrides={}
        for x in c.execute('SELECT product_id,channel,rate FROM commissions'): commission_overrides.setdefault(str(x[0]),{})[x[1]]=x[2]
        commissions={}
        for p in products: commissions[str(p['id'])]={**category_commissions.get(p['category'],{}),**commission_overrides.get(str(p['id']),{})}
        prices={}
        for x in c.execute('SELECT product_id,channel,price FROM prices'): prices.setdefault(str(x[0]),{})[x[1]]=x[2]
        raw_rates=[dict(x) for x in c.execute('SELECT desi,price FROM shipping_rates ORDER BY desi')]
        shipping_rates=[]
        for row in raw_rates:
            if shipping_rates and shipping_rates[-1]['price']==row['price'] and shipping_rates[-1]['max_desi']+1==row['desi']:
                shipping_rates[-1]['max_desi']=row['desi']
            else: shipping_rates.append({'min_desi':row['desi'],'max_desi':row['desi'],'price':row['price']})
        creds=load_credentials()
        last_sync=c.execute("SELECT * FROM sync_log ORDER BY id DESC LIMIT 1").fetchone()
        sentos_count=c.execute("SELECT COUNT(*) FROM sentos_products").fetchone()[0]
        product_summary={
            'total': c.execute('SELECT COUNT(*) FROM products').fetchone()[0],
            'matched': c.execute("SELECT COUNT(*) FROM products WHERE match_status='eslesti'").fetchone()[0],
            'unmatched': c.execute("SELECT COUNT(*) FROM products WHERE match_status IN ('eslesmedi','cakisiyor')").fetchone()[0],
            'waiting': c.execute("SELECT COUNT(*) FROM products WHERE match_status='bekliyor'").fetchone()[0],
        }
        now=datetime.now().isoformat(timespec='seconds')
        campaign_summary={
            'planned': c.execute("SELECT COUNT(*) FROM campaigns WHERE status='taslak'").fetchone()[0],
            'active': c.execute("SELECT COUNT(*) FROM campaigns WHERE status<>'taslak' AND starts<=? AND ends>=?",(now,now)).fetchone()[0],
            'upcoming': c.execute("SELECT COUNT(*) FROM campaigns WHERE status<>'taslak' AND starts>?",(now,)).fetchone()[0],
            'completed': c.execute("SELECT COUNT(*) FROM campaigns WHERE ends<?",(now,)).fetchone()[0],
        }
        price_summary={
            'pending': c.execute("SELECT COUNT(*) FROM price_update_items WHERE status IN ('taslak','bekliyor')").fetchone()[0],
            'failed': c.execute("SELECT COUNT(*) FROM price_update_items WHERE status='basarisiz'").fetchone()[0],
            'success': c.execute("SELECT COUNT(*) FROM price_update_items WHERE status='basarili'").fetchone()[0],
        }
        last_run=c.execute("SELECT * FROM price_update_runs ORDER BY id DESC LIMIT 1").fetchone()
    return {'products':products,'channels':channels,'offers':offers,'campaigns':campaigns,'commissions':commissions,'commission_overrides':commission_overrides,'category_commissions':category_commissions,'prices':prices,'shipping_rates':shipping_rates,'today':datetime.now().isoformat(),'product_summary':product_summary,'campaign_summary':campaign_summary,'price_summary':price_summary,'last_price_run':dict(last_run) if last_run else None,'sentos':{'configured':bool(creds),'base_url':creds['base_url'] if creds else '','username':creds['username'] if creds else '','product_count':sentos_count,'last_sync':dict(last_sync) if last_sync else None}}

class Handler(SimpleHTTPRequestHandler):
    def translate_path(self,path): return str(ROOT/'static'/urlparse(path).path.lstrip('/'))
    def end_headers(self):
        self.send_header('Cache-Control','no-store, no-cache, must-revalidate')
        return super().end_headers()
    def json(self,obj,status=200):
        data=json.dumps(obj,ensure_ascii=False).encode(); self.send_response(status); self.send_header('Content-Type','application/json; charset=utf-8'); self.send_header('Content-Length',str(len(data))); self.end_headers(); self.wfile.write(data)
    def do_GET(self):
        if self.path=='/api/data': return self.json(payload())
        if self.path=='/api/update/check':
            try:
                remote=urlopen(UPDATE_VERSION_URL,timeout=8).read().decode('utf-8').strip()
                return self.json({'ok':True,'local':VERSION,'remote':remote,'available':remote!=VERSION})
            except Exception as e:
                return self.json({'ok':False,'local':VERSION,'error':'Güncelleme sunucusuna ulaşılamadı.'})
        if self.path=='/': self.path='/index.html'
        return super().do_GET()
    def do_POST(self):
        try:
            n=int(self.headers.get('Content-Length',0)); body=json.loads(self.rfile.read(n) or b'{}')
            if self.path=='/api/update/install':
                if os.name != 'nt': raise ValueError('Otomatik güncelleme yalnızca Windows programında çalışır.')
                script=ROOT/'Guncelle.cmd'
                if not script.exists(): raise ValueError('Güncelleme dosyası bulunamadı.')
                subprocess.Popen(['cmd','/c',str(script)],cwd=str(ROOT),creationflags=0x00000008)
                threading.Timer(2,lambda:os._exit(0)).start()
                return self.json({'ok':True,'message':'Güncelleme indiriliyor. Program birazdan yeniden açılacak.'})
            if self.path=='/api/sentos/settings':
                base=str(body.get('base_url','')).strip(); user=str(body.get('username','')).strip(); password=str(body.get('password',''))
                old=load_credentials()
                if not password and old: password=old.get('password','')
                if not base.startswith('https://') or not user or not password: raise ValueError('HTTPS Sentos adresi, kullanıcı adı ve parola zorunludur.')
                save_credentials(base,user,password)
                return self.json({'ok':True,'message':'Sentos bilgileri bu Windows kullanıcısı için şifreli kaydedildi.'})
            if self.path=='/api/sentos/test':
                creds=load_credentials()
                if not creds: raise ValueError('Önce Sentos bağlantı bilgilerini kaydedin.')
                result=SentosClient(**creds).products(size=1,page=1)
                items=product_list(result)
                with db() as c: c.execute("INSERT INTO sync_log(kind,status,detail,item_count) VALUES('connection','ok','Salt okunur bağlantı testi başarılı',?)",(len(items),))
                return self.json({'ok':True,'message':'Bağlantı başarılı. Sentos ürün servisi cevap verdi.','sample_count':len(items)})
            if self.path=='/api/sentos/sync':
                creds=load_credentials()
                if not creds: raise ValueError('Önce Sentos bağlantı bilgilerini kaydedin.')
                client=SentosClient(**creds); all_items=[]; seen=set()
                for page in range(1,101):
                    batch=product_list(client.products(size=100,page=page))
                    fresh=[x for x in batch if isinstance(x,dict) and x.get('id') not in seen]
                    all_items.extend(fresh); seen.update(x.get('id') for x in fresh)
                    if len(batch)<100 or not fresh: break
                now=datetime.now().isoformat(timespec='seconds')
                linked=created=0
                with db() as c:
                    for x in all_items:
                        c.execute('''INSERT INTO sentos_products(sentos_id,sku,name,barcode,purchase_price,vat_rate,desi,prices_json,synced_at) VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(sentos_id) DO UPDATE SET sku=excluded.sku,name=excluded.name,barcode=excluded.barcode,purchase_price=excluded.purchase_price,vat_rate=excluded.vat_rate,desi=excluded.desi,prices_json=excluded.prices_json,synced_at=excluded.synced_at''',(x.get('id'),str(x.get('sku') or ''),str(x.get('name') or ''),str(x.get('barcode') or ''),num(x.get('purchase_price')),num(x.get('vat_rate')),num(x.get('volumetric_weight')),json.dumps(x.get('prices') or {},ensure_ascii=False),now))
                        sku=str(x.get('sku') or '').strip()
                        if sku:
                            existing=c.execute("SELECT id,desi FROM products WHERE upper(sku)=upper(?)",(sku,)).fetchone()
                            if existing:
                                # Ürün adı, barkod ve maliyet Sentos'un kaynağıdır. Yerel desi/kategori/web fiyatı korunur.
                                c.execute('''UPDATE products SET sentos_id=?,sentos_sku=?,name=?,barcode=?,cost=?,vat_rate=?,match_status='eslesti',match_note='SKU ile otomatik eşleşti',last_synced_at=? WHERE id=?''',(x.get('id'),sku,str(x.get('name') or ''),str(x.get('barcode') or ''),num(x.get('purchase_price')),num(x.get('vat_rate')),now,existing['id']))
                                if not existing['desi'] and num(x.get('volumetric_weight')):
                                    c.execute("UPDATE products SET desi=? WHERE id=?",(int(num(x.get('volumetric_weight'))),existing['id']))
                                c.execute("INSERT INTO product_events(product_id,event_type,detail) VALUES(?,?,?)",(existing['id'],'sentos_sync','SKU ile Sentos kaydı eşleştirildi'))
                                linked+=1
                            else:
                                c.execute('''INSERT INTO products(sku,sentos_id,sentos_sku,name,barcode,category,cost,web_price,shipping,min_margin,desi,vat_rate,match_status,match_note,last_synced_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(sku,x.get('id'),sku,str(x.get('name') or ''),str(x.get('barcode') or ''),'',num(x.get('purchase_price')),0,0,.02,int(num(x.get('volumetric_weight'))),num(x.get('vat_rate')),'eslesti','Sentos SKU ile yeni kayıt oluşturuldu',now))
                                new_id=c.execute("SELECT id FROM products WHERE sku=?",(sku,)).fetchone()[0]
                                c.execute("INSERT INTO product_events(product_id,event_type,detail) VALUES(?,?,?)",(new_id,'sentos_sync','Sentos’tan yeni ürün oluşturuldu'))
                                created+=1
                    c.execute("UPDATE products SET match_status='eslesmedi',match_note='Sentos eşitlemesinde bu SKU bulunamadı' WHERE sentos_id IS NULL")
                    detail=f'Salt okunur eşitleme: {linked} eşleşti, {created} yeni Sentos ürünü eklendi'
                    c.execute("INSERT INTO sync_log(kind,status,detail,item_count) VALUES('products','ok',?,?)",(detail,len(all_items)))
                return self.json({'ok':True,'message':f'{len(all_items)} Sentos ürünü alındı. {linked} SKU eşleşti, {created} yeni ürün eklendi. Fiyat gönderilmedi.','count':len(all_items),'matched':linked,'created':created})
            if self.path=='/api/campaigns':
                with db() as c: c.execute('''INSERT INTO campaigns(product_id,channel,normal_price,campaign_price,normal_rate,campaign_rate,starts,ends,seller_discount,marketplace_share) VALUES(?,?,?,?,?,?,?,?,?,?)''',tuple(body.get(k) for k in ['product_id','channel','normal_price','campaign_price','normal_rate','campaign_rate','starts','ends','seller_discount','marketplace_share']))
                return self.json({'ok':True})
            if self.path=='/api/channels':
                with db() as c: c.execute("UPDATE channels SET fixed_fee=?,extra_rate=? WHERE code=?",(body['fixed_fee'],body['extra_rate'],body['code']))
                return self.json({'ok':True})
            if self.path=='/api/prices':
                pid=int(body['product_id']); channel=str(body['channel']); price=num(body['price'])
                if channel not in CHANNELS or price<=0: raise ValueError('Geçerli bir kanal ve sıfırdan büyük fiyat girin.')
                with db() as c:
                    old=c.execute("SELECT price FROM prices WHERE product_id=? AND channel=?",(pid,channel)).fetchone()
                    c.execute("INSERT INTO price_history(product_id,channel,old_price,new_price) VALUES(?,?,?,?)",(pid,channel,old[0] if old else None,price))
                    c.execute("INSERT INTO prices(product_id,channel,price,source,updated_at) VALUES(?,?,?,'manuel',CURRENT_TIMESTAMP) ON CONFLICT(product_id,channel) DO UPDATE SET price=excluded.price,source='manuel',updated_at=CURRENT_TIMESTAMP",(pid,channel,price))
                return self.json({'ok':True,'mode':'demo','message':'Fiyat taslak olarak kaydedildi; Sentos’a gönderilmedi.'})
            if self.path=='/api/prices/bulk':
                pid=int(body['product_id']); entries=body.get('prices',{})
                with db() as c:
                    for channel,value in entries.items():
                        price=num(value)
                        if channel not in CHANNELS or price<=0: continue
                        old=c.execute("SELECT price FROM prices WHERE product_id=? AND channel=?",(pid,channel)).fetchone()
                        c.execute("INSERT INTO price_history(product_id,channel,old_price,new_price) VALUES(?,?,?,?)",(pid,channel,old[0] if old else None,price))
                        c.execute("INSERT INTO prices(product_id,channel,price,source,updated_at) VALUES(?,?,?,'otomatik',CURRENT_TIMESTAMP) ON CONFLICT(product_id,channel) DO UPDATE SET price=excluded.price,source='otomatik',updated_at=CURRENT_TIMESTAMP",(pid,channel,price))
                return self.json({'ok':True,'message':f'{len(entries)} kanal fiyatı taslak olarak kaydedildi.'})
            if self.path=='/api/bulk-update':
                items=body.get('items',[])
                if not items: raise ValueError('Güncellenecek ürün seçilmedi.')
                with db() as c:
                    for item in items:
                        pid=int(item['product_id']); web=num(item['web_price'])
                        if web<=0: continue
                        c.execute("UPDATE products SET web_price=? WHERE id=?",(web,pid))
                        values={'web':web,'perakende':web,**item.get('prices',{})}
                        for channel,value in values.items():
                            price=num(value)
                            if channel not in CHANNELS or price<=0: continue
                            old=c.execute("SELECT price FROM prices WHERE product_id=? AND channel=?",(pid,channel)).fetchone()
                            c.execute("INSERT INTO price_history(product_id,channel,old_price,new_price) VALUES(?,?,?,?)",(pid,channel,old[0] if old else None,price))
                            c.execute("INSERT INTO prices(product_id,channel,price,source,updated_at) VALUES(?,?,?,'toplu',CURRENT_TIMESTAMP) ON CONFLICT(product_id,channel) DO UPDATE SET price=excluded.price,source='toplu',updated_at=CURRENT_TIMESTAMP",(pid,channel,price))
                return self.json({'ok':True,'message':f'{len(items)} ürünün web ve pazaryeri fiyatları taslağa kaydedildi.'})
            if self.path=='/api/product':
                pid=int(body['product_id']); web=num(body['web_price']); cost=num(body['cost']); desi=max(0,int(num(body['desi']))); margin=num(body['min_margin'])
                with db() as c:
                    ship=c.execute("SELECT price FROM shipping_rates WHERE desi=?",(desi,)).fetchone(); shipping=ship[0] if ship else 0
                    c.execute("UPDATE products SET web_price=?,cost=?,desi=?,shipping=?,min_margin=? WHERE id=?",(web,cost,desi,shipping,margin,pid))
                    if web>0:
                        c.execute("INSERT INTO prices(product_id,channel,price,source,updated_at) VALUES(?, 'web',?,'manuel',CURRENT_TIMESTAMP) ON CONFLICT(product_id,channel) DO UPDATE SET price=excluded.price,source='manuel',updated_at=CURRENT_TIMESTAMP",(pid,web))
                        if body.get('sync_retail',True): c.execute("INSERT INTO prices(product_id,channel,price,source,updated_at) VALUES(?, 'perakende',?,'otomatik',CURRENT_TIMESTAMP) ON CONFLICT(product_id,channel) DO UPDATE SET price=excluded.price,source='otomatik',updated_at=CURRENT_TIMESTAMP",(pid,web))
                return self.json({'ok':True})
            if self.path=='/api/commissions':
                pid=int(body['product_id'])
                with db() as c:
                    for channel,value in body.get('rates',{}).items():
                        if channel in CHANNELS: c.execute("INSERT OR REPLACE INTO commissions VALUES(?,?,?)",(pid,channel,num(value)))
                return self.json({'ok':True})
            if self.path=='/api/category-commissions':
                category=str(body['category'])
                with db() as c:
                    for channel,value in body.get('rates',{}).items():
                        if channel in CHANNELS and channel not in ('web','perakende'): c.execute("INSERT INTO category_commissions(category,channel,rate,updated_at) VALUES(?,?,?,CURRENT_TIMESTAMP) ON CONFLICT(category,channel) DO UPDATE SET rate=excluded.rate,updated_at=CURRENT_TIMESTAMP",(category,channel,num(value)))
                return self.json({'ok':True,'message':f'{category} kategori komisyonları güncellendi.'})
            if self.path=='/api/product-commissions':
                pid=int(body['product_id'])
                with db() as c:
                    for channel,value in body.get('rates',{}).items():
                        if channel not in CHANNELS or channel in ('web','perakende'): continue
                        if value is None or value=='': c.execute("DELETE FROM commissions WHERE product_id=? AND channel=?",(pid,channel))
                        else: c.execute("INSERT OR REPLACE INTO commissions(product_id,channel,rate) VALUES(?,?,?)",(pid,channel,num(value)))
                return self.json({'ok':True,'message':'Ürüne özel komisyon kuralları güncellendi.'})
            if self.path=='/api/shipping':
                low=max(0,int(num(body['min_desi']))); high=max(low,int(num(body['max_desi']))); price=num(body['price'])
                if price<0 or high>500: raise ValueError('Kargo tarifesi veya desi aralığı geçersiz.')
                with db() as c:
                    for desi in range(low,high+1): c.execute("INSERT OR REPLACE INTO shipping_rates(desi,price,updated_at) VALUES(?,?,CURRENT_TIMESTAMP)",(desi,price))
                    c.execute("UPDATE products SET shipping=? WHERE desi BETWEEN ? AND ?",(price,low,high))
                return self.json({'ok':True,'message':f'{low}–{high} desi kargo bedeli güncellendi.'})
            self.json({'error':'Bulunamadı'},404)
        except Exception as e: self.json({'error':str(e)},400)

def bootstrap():
    init_db()
    upload=ROOT.parent/'upload'
    if not db().execute("SELECT 1 FROM products LIMIT 1").fetchone(): import_master(upload/'2024 Haziran Fiyat Taslak.xlsx')
    elif (not db().execute("SELECT 1 FROM prices LIMIT 1").fetchone() or not db().execute("SELECT 1 FROM shipping_rates LIMIT 1").fetchone()) and (upload/'2024 Haziran Fiyat Taslak.xlsx').exists(): import_master(upload/'2024 Haziran Fiyat Taslak.xlsx')
    if not db().execute("SELECT 1 FROM offers LIMIT 1").fetchone():
        import_offers(upload/'trendyol.xlsx'); import_offers(upload/'n11.xls')
    normalize_category_commissions()

if __name__=='__main__':
    bootstrap(); url='http://127.0.0.1:8765'
    print(f'Kochler Fiyat Otomasyonu v0.5: {url}\nKapatmak için Ctrl+C')
    threading.Timer(1,lambda:webbrowser.open(url)).start()
    ThreadingHTTPServer(('127.0.0.1',8765),Handler).serve_forever()
